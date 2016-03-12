# -*- coding: cp1252 -*-

#  -- Funciones para crear y modificar la Ranked Statistics Spreadsheet --

from openpyxl import Workbook
from RiotAPIApp import *

def createSpreadsheet(summonerData, APIKey):#       Para crear archivo xlsx con info de Invocador  ***Faltan cosas
        wb = Workbook()
        ws = wb.active#         Seleccionas la hoja activa por default se crea una
       #     Para crear la primer hoja que tendra datos generales del Invocador
        ws.title = "Summoner Info"
        ws.row_dimensions[1].hidden = True
        ws.column_dimensions['A'].width = 20
        ws['A1'] = 0# Agregamos el Revision Date para validar actualizaciones
        ws['A2'] = "Summoner ID:"
        ws['A3'] = "Summoner Name:"
        ws['A4'] = "Region:"
        ws['A5'] = "Icon:"
        ws['A6'] = "Level:"
        ws['A7'] = "Rank/Division:"
        ws['A8'] = "League Points:"
        ws['A9'] = "Win/Loss:"
        ws['A10'] = "KDA:"
        ws['A11'] = "Average Creep Score:"
        ws['D3'] = "My last 5 Games"
        ws['D4'] = "Type"
        ws['E4'] = "Result"
        ws['F4'] = "Champ"
        ws['G4'] = "KDA"
        ws['H4'] = "CS/min"

        # Segunta hoja con datos de partidas ranked 
        ws1 = wb.create_sheet()
        ws1.title = "Ranked Games"
        ws = wb['Ranked Games']
        ws['A2'] = "GameID"
        ws ['A3'] = 0 #Asignamos un valor 0 a primer matchID para actualizar
        ws.column_dimensions['A'].hidden = True
        ws['B2'] = "Game"
        ws['C2'] = "Win/Loss"
        ws['D2'] = "Champ Played"
        ws['E2'] = "Summoner spells"
        ws['F2'] = "Role"
        ws['G2'] = "Kills"
        ws['H2'] = "Deaths"
        ws['I2'] = "Assists"
        ws['J2'] = "Kill participation"
        ws['K2'] = "Cs per min"
        ws['L2'] = "Total gold"
        ws['M2'] = "Comments"
        
        wb.save('RankedData.xlsx')
        print "\nSpreadsheet created\n"

def spreadsheetUpdater(summonerData, APIKey):# Actualiza el spreadsheet *****Faltan muchas cosas
        summonerName = summonerData.keys()[0]
        ID = (str)(summonerData[summonerName]['id'])
        region = summonerData[summonerName]['region']
        summonerRankedData = requestsEngine.requestRankedData(summonerData, APIKey)
        staticChampion = staticRequests.requestStaticChampion(summonerData, APIKey)
        staticItem = staticRequests.requestStaticItem(summonerData, APIKey)
        staticSpells = staticRequests.requestStaticSpells(summonerData, APIKey)
        
        from openpyxl import load_workbook
        # Actualizar la hoja de Summoner Info
        wb = load_workbook('RankedData.xlsx')
        ws = wb["Summoner Info"]
        ws['B2'] = summonerData[summonerName]['id']
        ws['B3'] = summonerData[summonerName]['name']
        ws['B4'] = str.capitalize(summonerData[summonerName]['region'])
        ws['B5'] = summonerData[summonerName]['profileIconId']
        ws['B6'] = summonerData[summonerName]['summonerLevel']
        ws['B7'] = summonerRankedData[ID][0]['tier'] + " " + summonerRankedData[ID][0]['entries'][0]['division']
        ws['B8'] = summonerRankedData[ID][0]['entries'][0]['leaguePoints']

        # Actualizar los ultimos 5 matches
        summonerRecentGames = requestsEngine.requestRecentGames(summonerData, APIKey)
        i = 0
        a = 5
        while i < 5:#Loop para llenar los datos de los 5 matches **** No me deja definir variables dentro del loop ** No puedo sumar para CS
                a = str(a)
                ws['D'+ a] = summonerRecentGames['games'][i]['subType']
                if summonerRecentGames['games'][i]['stats']['win']:
                        ws['E'+ a] = "Win"
                else:
                        ws['E' + a] = "Loss"
                ws['F' + a] = staticChampion['data'][(str)(summonerRecentGames['games'][i]['championId'])]['name']# Campeon usado
                
                # Para poner las kills, deaths y assists primero hay que buscar si el Key existe en el dictionary
                if 'championsKilled' in summonerRecentGames['games'][i]['stats']:
                        kills = (str)(summonerRecentGames['games'][i]['stats']['championsKilled'])
                else:
                        kills = "0"
                if 'numDeaths' in summonerRecentGames['games'][i]['stats']:
                        deaths = (str)(summonerRecentGames['games'][i]['stats']['numDeaths'])
                else:
                        deaths = "0"
                if 'assists' in summonerRecentGames['games'][i]['stats']:
                        assists = (str)(summonerRecentGames['games'][i]['stats']['assists'])
                else:
                        assists = "0"
                ws['G' + a] = kills + "/" + deaths + "/" + assists

                # Para poner CS primero se busan y se suman los cs de jungla (tuya y enemiga) y minions
                if 'neutralMinionsKilled' in summonerRecentGames['games'][i]['stats']:
                        jungla = summonerRecentGames['games'][i]['stats']['neutralMinionsKilled']
                else:
                        jungla = 0
                if 'minionsKilled' in summonerRecentGames['games'][i]['stats']:
                        minions = summonerRecentGames['games'][i]['stats']['minionsKilled']
                else:
                        minions = 0
                CS = jungla + minions
                # Calcular los minutos jugados
                if 'timePlayed' in summonerRecentGames['games'][i]['stats']:
                        timeMin = summonerRecentGames['games'][i]['stats']['timePlayed'] / 60.0
                else:
                        timeMin = 0
                ws['H' + a] = format((CS / timeMin), '.2f')
                
                a = int(a)
                i += 1
                a += 1
        # Actualizar Ranked Matchlist
        ws = wb['Ranked Games']
                
        a = 3 # ****Debe cambiar con respecto a la utima linea con info en la hoja (ws)
        a = str(a)
                
        matchData = requestsEngine.requetsMatchData(summonerData, APIKey) # Se llama para generar Data

        ### *** Todo esto debe estar en un loop *** ####
        
        ws['A' + a ] = matchData['matchId']# Llenamos el match ID de la partida

        matchDuration = matchData['matchDuration'] / 60.0#Cuanto duró el match en minutos
        
        ## Para calcular las muertes de los equipos. Para kill participation
        team100Deaths = 0
        team200Deaths = 0
        for i in matchData['participants']:
                if i['teamId'] == 100:
                        team100Deaths = i['stats']['deaths'] + team100Deaths
                else:
                        team200Deaths = i['stats']['deaths'] + team200Deaths
            
        for i in matchData['participantIdentities']:# Buscamos el summoner ID entre los participantes para obtener el usuario
            var2 = (int)(summonerData[summonerName]['id'])
            if i['player']['summonerId'] == var2:
                participantId = i['participantId']# Jalamos el id de participante de la partida
                break
        for i in matchData['participants']:
            if i['participantId'] == participantId: # Cuando encontramos el player id empezamos a llenar
                # Empezamos a llenar los datos
                # Win Loss
                if i['stats']['winner']:
                        ws['C' + a] = "Win"
                else:
                        ws['C' + a] = "Loss"

                ws['D' + a] = staticChampion['data'][(str)(i['championId'])]['name']
                # Summoner Spells
                spell1 = staticSpells['data'][(str)(i['spell1Id'])]['name']
                spell2 = staticSpells['data'][(str)(i['spell2Id'])]['name']
                ws['E' + a] = spell1 + " / " + spell2

                ws['F' + a] = i['timeline']['lane']
                ws['G' + a] = i['stats']['kills']
                ws['H' + a] = i['stats']['deaths']
                ws['I' + a] = i['stats']['assists']

                #Calcular kill participation
                killsAssists = (int)(i['stats']['kills']) + (int)(i['stats']['assists'])
                
                if i['teamId'] == 100:
                        ws['J' + a] = format((float(killsAssists) / team200Deaths), '.2f')
                else:
                        ws['J' + a] = format((float(killsAssists) / team100Deaths), '.2f')
                        
                #Se suman minions neutrales + minions para CS
                minionsCS = i['stats']['minionsKilled']
                jungleCS = i['stats']['neutralMinionsKilled']
                totalCS = minionsCS + jungleCS
                ws['K' + a] = format((totalCS / matchDuration), '.2f')# Minions per min
                
                # Total de oro
                ws['L' + a] = i['stats']['goldEarned']

                #Color de las celdas corresponde al color del equipo del participante
                
                break# Fin del loop de llenadera
        
        
        # Se cambia el revisionDate para validar futuras actualizaciones
        wb["Summoner Info"]['A1'] = summonerData[summonerName]['revisionDate']
        # Se guardan los cambios
        wb.save('RankedData.xlsx')
        print "\nYour Ranked Statistics Spreadsheet has been updated\n"
